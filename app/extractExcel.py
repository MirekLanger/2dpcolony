import openpyxl
import random

def readParameters(wb):
    parametersSheet = wb['Parameters']
    envRows = int(parametersSheet.cell(row=1, column=2).value)
    envColums = int(parametersSheet.cell(row=2, column=2).value)
    envRules = int(parametersSheet.cell(row=3, column=2).value)
    steps = int(parametersSheet.cell(row=4, column=2).value)
    return {
        'envRows': envRows,
        'envColumns': envColums,
        'envRules': envRules,
        'steps': steps
    }

def readEnvironment(wb, rows, columns, rules):
    environmentSheet = wb['Environment']
    envMatrix = [['e' for j in range(columns)] for i in range(rows)]
    for i in range(1, rows + 1):
        for j in range(1, columns + 1):
            envMatrix[i-1][j-1] = environmentSheet.cell(row=i, column=j).value.split(',')
    envRulesSheet = wb['Environment_rules']
    envRules = []
    for i in range(1, rules + 1):
        rule = {
            'left': envRulesSheet.cell(row=i, column=1).value.split(','),
            'right': envRulesSheet.cell(row=i, column=3).value.split(',')
        }
        envRules.append(rule)
    return {
        'matrix': envMatrix,
        'rules': envRules
    }

def readAgents(wb):
    agentsSheet  = wb['Agents']
    agents = []
    lineIndex = 1
    while agentsSheet.cell(row=lineIndex, column=1).value != 'AgentsEnd':
        #Definition of an agent begins
        if agentsSheet.cell(row=lineIndex, column=1).value == 'AgentBegin':
            agent = {}
            programs = []
            agent['id'] = agentsSheet.cell(row=lineIndex, column=3).value
            agent['contents'] = agentsSheet.cell(row=lineIndex, column=5).value.split(',')
            agent['coordinates'] = {
                'i': int(agentsSheet.cell(row=lineIndex, column=7).value),
                'j': int(agentsSheet.cell(row=lineIndex, column=9).value)
            }
            agent['copies'] = int(agentsSheet.cell(row=lineIndex, column=11).value)
        #definition of a program begins
        if agentsSheet.cell(row=lineIndex, column=2).value == 'programBegin':
            program = []
        #definition of a rule
        if agentsSheet.cell(row=lineIndex, column=2).value == 'rule':
            rule = {
                'left': '',
                'operator': agentsSheet.cell(row=lineIndex, column=4).value,
                'right':agentsSheet.cell(row=lineIndex, column=5).value
            }
            if rule['operator'] in ['u', 'd', 'l', 'r']:
                rule['left'] = agentsSheet.cell(row=lineIndex, column=3).value.split(',')
            else:
                rule['left'] = agentsSheet.cell(row=lineIndex, column=3).value
            program.append(rule)
        # definition of a program ends
        if agentsSheet.cell(row=lineIndex, column=2).value == 'programEnd':
            programs.append(program)
        #definition of an agent ends
        if agentsSheet.cell(row=lineIndex, column=1).value == 'AgentEnd':
            agent['programs'] = programs
            if agent['copies'] > 2:
                id = agent['id']
                for i in range (agent['copies']):
                    agent['id'] += id + '_' + str(i)
                    print(agent['id'])
                    agents.append(agent)
            else:
                agents.append(agent)
        lineIndex += 1
    return agents

def getColonie(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    parameters = readParameters(wb)
    environment = readEnvironment(wb, parameters['envRows'], parameters['envColumns'], parameters['envRules'])
    agents = readAgents(wb)
    return {
        'parameters': parameters,
        'environment': environment,
        'agents': agents
    }





